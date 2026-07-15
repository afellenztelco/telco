import io
import os
import shutil
import enum
import secrets
import re
from datetime import date, datetime
from typing import List, Optional
from fastapi import FastAPI, Depends, Request, HTTPException, status, Cookie, Form, UploadFile, File
from fastapi.responses import StreamingResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from sqlalchemy import create_engine, Column, Integer, String, Date, Numeric, DateTime, desc, Boolean, text
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from pydantic import BaseModel
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

os.makedirs("adjuntos", exist_ok=True)
DATABASE_URL = "sqlite:///./enlaces.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

class Usuario(Base):
    __tablename__ = "usuarios"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(50), unique=True, index=True, nullable=False)
    password = Column(String(100), nullable=False)

class EstadoEnlace(str, enum.Enum):
    ACTIVO = "ACTIVO"
    INACTIVO = "INACTIVO"

class TipoEnlace(str, enum.Enum):
    FIBRA_DEDICADO = "Fibra Optica dedicado"
    FIBRA_ASIMETRICO = "Fibra Optica asimetrico"
    RF = "Radio Frecuencia"
    LEO = "LEO"
    LEO_ITINERANTE = "LEO Itinerante"
    LEO_1TB = "LEO 1 TB"
    GEO = "GEO"
    L2L = "L2L (transporte de datos)"

class Tarifa(Base):
    __tablename__ = "tarifas"
    id = Column(Integer, primary_key=True, index=True)
    tipo_enlace = Column(String(50), nullable=False)
    ancho_banda = Column(String(50), nullable=True, default="N/A") 
    moneda = Column(String(10), nullable=False, default="USD")
    precio_costo = Column(Numeric(12, 2), default=0.00)
    costo_mantenimiento = Column(Numeric(12, 2), default=0.00)
    precio_venta_sin_iva = Column(Numeric(12, 2), default=0.00)
    precio_venta_con_iva = Column(Numeric(12, 2), default=0.00)
    costo_instalacion = Column(Numeric(12, 2), default=0.00)

class ReporteFacturacion(Base):
    __tablename__ = "reportes_facturacion"
    id = Column(Integer, primary_key=True, index=True)
    fecha_generacion = Column(DateTime, default=datetime.now)
    archivo = Column(String(255), nullable=False)
    usuario = Column(String(50), nullable=False)
    comentario = Column(String(500), nullable=True)

class Enlace(Base):
    __tablename__ = "enlaces_telecom"
    id = Column(Integer, primary_key=True, index=True)
    referencia = Column(String(50), unique=True, index=True, nullable=False)
    organismo = Column(String(150), nullable=False)
    ubicacion = Column(String(150), nullable=True) 
    observaciones = Column(String(1000), nullable=True)
    archivo_adjunto = Column(String(1000), nullable=True)
    localidad = Column(String(100), nullable=False)
    tipo_enlace = Column(String(50), nullable=False)
    estado = Column(String(20), default="ACTIVO")
    ancho_banda = Column(String(50), nullable=False)
    moneda = Column(String(10), nullable=False, default="USD")
    precio_costo = Column(Numeric(12, 2), default=0.00)
    costo_mantenimiento = Column(Numeric(12, 2), default=0.00)
    precio_venta_sin_iva = Column(Numeric(12, 2), default=0.00)
    precio_venta_con_iva = Column(Numeric(12, 2), default=0.00)
    costo_instalacion = Column(Numeric(12, 2), default=0.00)
    fecha_alta = Column(Date, nullable=False)
    proveedor = Column(String(100), nullable=False)
    cupo_transferencia = Column(String(100), nullable=True)
    coordenadas_gps = Column(String(100), nullable=True)
    sn_antena = Column(String(100), nullable=True)
    sn_modem = Column(String(100), nullable=True)
    nro_te = Column(String(50), nullable=True)
    nro_item = Column(String(50), nullable=True)
    eliminado = Column(Boolean, default=False)

class LogActividad(Base):
    __tablename__ = "logs_actividad"
    id = Column(Integer, primary_key=True, index=True)
    fecha_hora = Column(DateTime, default=datetime.now)
    usuario = Column(String(50), nullable=False)
    accion = Column(String(50), nullable=False)
    entidad_id = Column(Integer, nullable=False)
    detalle = Column(String(500), nullable=False)

Base.metadata.create_all(bind=engine)

db_init = SessionLocal()
if db_init.query(Usuario).count() == 0:
    db_init.add(Usuario(username="scortes", password="acatenestusistemita"))
    db_init.add(Usuario(username="afellenz", password="Va2005fe"))
    db_init.commit()
db_init.close()

with engine.connect() as conn:
    try: conn.execute(text("ALTER TABLE enlaces_telecom ADD COLUMN archivo_adjunto VARCHAR(1000)")); conn.commit()
    except Exception: pass 
    try: conn.execute(text("ALTER TABLE enlaces_telecom ADD COLUMN eliminado BOOLEAN DEFAULT 0")); conn.commit()
    except Exception: pass
    try: conn.execute(text("ALTER TABLE reportes_facturacion ADD COLUMN comentario VARCHAR(500)")); conn.commit()
    except Exception: pass

class UsuarioCreate(BaseModel):
    username: str
    password: str

class TarifaCreate(BaseModel):
    tipo_enlace: str
    moneda: Optional[str] = "USD"
    precio_costo: float
    costo_mantenimiento: float
    precio_venta_sin_iva: float
    precio_venta_con_iva: float
    costo_instalacion: float

class EnlaceCreate(BaseModel):
    referencia: str
    organismo: str
    ubicacion: Optional[str] = None 
    observaciones: Optional[str] = None
    localidad: str
    tipo_enlace: TipoEnlace
    ancho_banda: str
    moneda: Optional[str] = "USD"
    precio_costo: float
    costo_mantenimiento: float
    precio_venta_sin_iva: float
    precio_venta_con_iva: float
    costo_instalacion: float
    fecha_alta: date
    proveedor: str
    cupo_transferencia: Optional[str] = None
    coordenadas_gps: Optional[str] = None
    sn_antena: Optional[str] = None
    sn_modem: Optional[str] = None
    nro_te: Optional[str] = None
    nro_item: Optional[str] = None

class FacturacionCreate(BaseModel):
    comentario: Optional[str] = ""

app = FastAPI()
app.mount("/adjuntos", StaticFiles(directory="adjuntos"), name="adjuntos")
templates = Jinja2Templates(directory="templates")

def get_db():
    db = SessionLocal()
    try: yield db
    finally: db.close()

def verificar_usuario(session_user: str = Cookie(None), db: Session = Depends(get_db)):
    if not session_user: raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Inicie sesión")
    user = db.query(Usuario).filter(Usuario.username == session_user).first()
    if not user: raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Inicie sesión")
    return session_user

def extract_numeric(text):
    match = re.search(r'[\d\.]+', str(text))
    if match:
        try: return float(match.group(0))
        except: return 1.0
    return 1.0

# NUEVA FUNCION PARA FORMATEAR A 1.024,00
def format_currency(value):
    try:
        val = float(value)
        # Transforma a formato 1,234.56
        formateado = f"{val:,.2f}"
        # Intercambia comas por puntos y viceversa
        return formateado.replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return "0,00"

def aplicar_tarifa_a_enlaces(tarifa, db, user):
    enlaces = db.query(Enlace).filter(
        Enlace.tipo_enlace == tarifa.tipo_enlace,
        (Enlace.eliminado == False) | (Enlace.eliminado.is_(None))
    ).all()
    count = 0
    for e in enlaces:
        base_sin_iva = float(tarifa.precio_venta_sin_iva)
        multiplier = extract_numeric(e.ancho_banda) if tarifa.tipo_enlace in ["Fibra Optica dedicado", "L2L (transporte de datos)"] else 1.0
        e.moneda = "USD"
        e.precio_costo = tarifa.precio_costo
        e.costo_mantenimiento = tarifa.costo_mantenimiento
        e.costo_instalacion = tarifa.costo_instalacion
        e.precio_venta_sin_iva = base_sin_iva * multiplier
        e.precio_venta_con_iva = round(e.precio_venta_sin_iva * 1.21, 2)
        count += 1
    if count > 0:
        db.add(LogActividad(usuario=user, accion="ACTUALIZACION_MASIVA", entidad_id=tarifa.id, detalle=f"Actualizados {count} enlaces por cambio en Tarifario ({tarifa.tipo_enlace})"))
    return count

def enlace_to_dict(e):
    costo_inst_original = float(e.costo_instalacion or 0)
    costo_inst_cobrar = costo_inst_original
    if e.fecha_alta:
        hoy = date.today()
        try:
            if e.fecha_alta.year != hoy.year or e.fecha_alta.month != hoy.month: costo_inst_cobrar = 0.0
        except AttributeError:
            try:
                fecha_obj = datetime.strptime(str(e.fecha_alta).split()[0], "%Y-%m-%d").date()
                if fecha_obj.year != hoy.year or fecha_obj.month != hoy.month: costo_inst_cobrar = 0.0
            except: pass
            
    mantenimiento_base = float(e.costo_mantenimiento or 0)

    return {
        "id": e.id,
        "referencia": e.referencia,
        "organismo": e.organismo,
        "ubicacion": e.ubicacion,
        "observaciones": e.observaciones,
        "archivo_adjunto": e.archivo_adjunto,
        "localidad": e.localidad,
        "tipo_enlace": str(e.tipo_enlace),
        "estado": str(e.estado),
        "ancho_banda": e.ancho_banda,
        "moneda": "USD",
        "precio_costo": float(e.precio_costo or 0),
        "costo_mantenimiento": mantenimiento_base,
        "costo_mantenimiento_con_iva": round(mantenimiento_base * 1.21, 2),
        "precio_venta_sin_iva": float(e.precio_venta_sin_iva or 0),
        "precio_venta_con_iva": float(e.precio_venta_con_iva or 0),
        "costo_instalacion": costo_inst_cobrar, 
        "costo_instalacion_original": costo_inst_original, 
        "fecha_alta": str(e.fecha_alta) if e.fecha_alta else "",
        "proveedor": e.proveedor,
        "cupo_transferencia": e.cupo_transferencia,
        "coordenadas_gps": e.coordenadas_gps,
        "sn_antena": e.sn_antena,
        "sn_modem": e.sn_modem,
        "nro_te": e.nro_te,
        "nro_item": e.nro_item
    }

@app.get("/")
def pagina_bienvenida(request: Request, session_user: str = Cookie(None), db: Session = Depends(get_db)):
    if session_user:
        if db.query(Usuario).filter(Usuario.username == session_user).first():
            return RedirectResponse(url="/sistema", status_code=status.HTTP_303_SEE_OTHER)
    return templates.TemplateResponse(request=request, name="login.html")

@app.post("/login")
def procesar_login(username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(Usuario).filter(Usuario.username == username).first()
    if user and secrets.compare_digest(password, user.password):
        response = RedirectResponse(url="/sistema", status_code=status.HTTP_303_SEE_OTHER)
        response.set_cookie(key="session_user", value=username, httponly=True)
        return response
    return RedirectResponse(url="/?error=true", status_code=status.HTTP_303_SEE_OTHER)

@app.get("/sistema")
def leer_interfaz(request: Request, session_user: str = Cookie(None), db: Session = Depends(get_db)):
    if not session_user or not db.query(Usuario).filter(Usuario.username == session_user).first():
        return RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)
    return templates.TemplateResponse(request=request, name="index.html", context={"session_user": session_user})

@app.get("/logout")
def logout():
    response = RedirectResponse(url="/", status_code=status.HTTP_303_SEE_OTHER)
    response.delete_cookie("session_user")
    return response

@app.get("/tarifas/")
def listar_tarifas(db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    return db.query(Tarifa).all()

@app.post("/tarifas/")
def crear_tarifa(t: TarifaCreate, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    if db.query(Tarifa).filter(Tarifa.tipo_enlace == t.tipo_enlace).first():
        raise HTTPException(status_code=400, detail="Ya existe una tarifa. Edite la existente.")
    nueva_tarifa = Tarifa(**t.model_dump(), ancho_banda="N/A")
    db.add(nueva_tarifa)
    db.commit()
    db.refresh(nueva_tarifa)
    aplicar_tarifa_a_enlaces(nueva_tarifa, db, user)
    db.commit()
    return {"ok": True}

@app.put("/tarifas/{id}")
def editar_tarifa(id: int, t: TarifaCreate, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    db_t = db.query(Tarifa).filter(Tarifa.id == id).first()
    if db_t:
        for key, value in t.model_dump().items(): setattr(db_t, key, value)
        db.commit()
        aplicar_tarifa_a_enlaces(db_t, db, user)
        db.commit()
    return {"ok": True}

@app.delete("/tarifas/{id}")
def eliminar_tarifa(id: int, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    db_t = db.query(Tarifa).filter(Tarifa.id == id).first()
    if db_t:
        db.delete(db_t)
        db.commit()
    return {"ok": True}

@app.get("/usuarios/")
def listar_usuarios(db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    if user != "afellenz": raise HTTPException(status_code=403, detail="No autorizado")
    return [{"id": u.id, "username": u.username} for u in db.query(Usuario).all()]

@app.post("/usuarios/")
def crear_usuario(u: UsuarioCreate, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    if user != "afellenz": raise HTTPException(status_code=403, detail="No autorizado")
    db.add(Usuario(username=u.username, password=u.password))
    db.commit()
    return {"ok": True}

@app.put("/usuarios/{id}")
def editar_usuario(id: int, u: UsuarioCreate, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    if user != "afellenz": raise HTTPException(status_code=403, detail="No autorizado")
    db_u = db.query(Usuario).filter(Usuario.id == id).first()
    if db_u:
        db_u.username = u.username; db_u.password = u.password; db.commit()
    return {"ok": True}

@app.delete("/usuarios/{id}")
def eliminar_usuario(id: int, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    if user != "afellenz": raise HTTPException(status_code=403, detail="No autorizado")
    db_u = db.query(Usuario).filter(Usuario.id == id).first()
    if db_u: db.delete(db_u); db.commit()
    return {"ok": True}

@app.post("/enlaces/")
def crear_enlace(enlace: EnlaceCreate, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    enlace_data = enlace.model_dump()
    enlace_data['tipo_enlace'] = enlace.tipo_enlace.value
    nuevo_enlace = Enlace(**enlace_data)
    db.add(nuevo_enlace)
    db.commit()
    db.add(LogActividad(usuario=user, accion="CREAR", entidad_id=nuevo_enlace.id, detalle=f"Creación: {nuevo_enlace.referencia}"))
    db.commit()
    return {"ok": True, "id": nuevo_enlace.id}

@app.get("/enlaces/{id}")
def obtener_enlace(id: int, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    e = db.query(Enlace).filter(Enlace.id == id).first()
    if e: return enlace_to_dict(e)
    raise HTTPException(status_code=404)

@app.put("/enlaces/{id}")
def actualizar_enlace(id: int, e: EnlaceCreate, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    db_e = db.query(Enlace).filter(Enlace.id == id).first()
    cambios = []
    e_data = e.model_dump()
    e_data['tipo_enlace'] = e.tipo_enlace.value
    for key, value in e_data.items():
        if str(getattr(db_e, key)) != str(value): cambios.append(f"{key}: {getattr(db_e, key)} -> {value}")
        setattr(db_e, key, value)
    if cambios:
        db.add(LogActividad(usuario=user, accion="EDITAR", entidad_id=id, detalle=" | ".join(cambios)[:500]))
    db.commit()
    return {"ok": True, "id": id}

@app.post("/enlaces/{id}/adjunto")
def subir_adjunto(id: int, files: List[UploadFile] = File(...), db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    enlace = db.query(Enlace).filter(Enlace.id == id).first()
    if not enlace: raise HTTPException(status_code=404)
    if len(files) > 10: raise HTTPException(status_code=400, detail="Máximo 10 archivos.")
    nombres_archivos = enlace.archivo_adjunto.split(',') if enlace.archivo_adjunto else []
    for file in files:
        if file.filename:
            filename = f"id{id}_{secrets.token_hex(2)}_{file.filename.replace(' ', '_').replace(',', '')}"
            with open(os.path.join("adjuntos", filename), "wb") as buffer: shutil.copyfileobj(file.file, buffer)
            nombres_archivos.append(filename)
    enlace.archivo_adjunto = ",".join(nombres_archivos)
    db.add(LogActividad(usuario=user, accion="ADJUNTAR", entidad_id=id, detalle=f"Se subieron {len(files)} archivos."))
    db.commit()
    return {"ok": True, "filenames": nombres_archivos}

@app.delete("/enlaces/{id}/adjunto/{filename:path}")
def eliminar_adjunto(id: int, filename: str, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    enlace = db.query(Enlace).filter(Enlace.id == id).first()
    if not enlace or not enlace.archivo_adjunto: raise HTTPException(status_code=404)
    archivos = enlace.archivo_adjunto.split(',')
    if filename in archivos:
        archivos.remove(filename)
        enlace.archivo_adjunto = ",".join(archivos) if archivos else None
        if os.path.exists(os.path.join("adjuntos", filename)):
            try: os.remove(os.path.join("adjuntos", filename))
            except: pass
        db.add(LogActividad(usuario=user, accion="ELIMINAR_ADJUNTO", entidad_id=enlace.id, detalle=f"Archivo borrado: {filename}"))
        db.commit()
        return {"ok": True}
    raise HTTPException(status_code=404, detail="Archivo no encontrado")

@app.get("/enlaces/")
def listar_enlaces(db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    enlaces = db.query(Enlace).all()
    resultado = []
    for e in enlaces:
        try:
            if e.eliminado not in [True, 1, "1", "true", "True"]: resultado.append(enlace_to_dict(e))
        except: pass
    return resultado

@app.patch("/enlaces/{enlace_id}/estado")
def cambiar_estado(enlace_id: int, estado: EstadoEnlace, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    enlace = db.query(Enlace).filter(Enlace.id == enlace_id).first()
    if enlace.estado != estado:
        db.add(LogActividad(usuario=user, accion="CAMBIAR_ESTADO", entidad_id=enlace.id, detalle=f"Estado: {enlace.estado} -> {estado.value}"))
        enlace.estado = estado.value
        db.commit()
    return {"ok": True}

@app.patch("/enlaces/{id}/eliminar")
def eliminar_enlace(id: int, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    enlace = db.query(Enlace).filter(Enlace.id == id).first()
    enlace.eliminado = True
    db.add(LogActividad(usuario=user, accion="ELIMINAR", entidad_id=enlace.id, detalle=f"Eliminación lógica: {enlace.referencia}"))
    db.commit()
    return {"ok": True}

@app.get("/logs/")
def listar_logs(db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    return db.query(LogActividad).order_by(desc(LogActividad.fecha_hora)).limit(100).all()

@app.get("/facturaciones/")
def listar_facturaciones(db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    reps = db.query(ReporteFacturacion).order_by(desc(ReporteFacturacion.fecha_generacion)).all()
    return [{"id": r.id, "fecha_generacion": r.fecha_generacion.isoformat(), "archivo": r.archivo, "usuario": r.usuario, "comentario": r.comentario} for r in reps]

@app.post("/facturaciones/generar")
def generar_facturacion(f: FacturacionCreate, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    enlaces = db.query(Enlace).all()
    data = [enlace_to_dict(e) for e in enlaces if e.eliminado not in [True, 1, "1", "true", "True"]]
    df = pd.DataFrame(data)
    
    filename = f"Facturacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{secrets.token_hex(2)}.pdf"
    filepath = os.path.join("adjuntos", filename)
    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), leftMargin=15, rightMargin=15, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    p_style = ParagraphStyle('WrapStyle', parent=styles['Normal'], fontSize=6, leading=7, alignment=TA_CENTER)
    p_style_ref = ParagraphStyle('RefStyle', parent=styles['Normal'], fontSize=5, leading=6, alignment=TA_CENTER)
    
    table_data = [['IT', 'Ref', 'Organismo', 'Ubicación', 'Localidad', 'Tipo', 'Ancho B.', 'Vta s/IVA', 'Vta c/IVA', 'Instal.', 'Mant.', 'Alta']]
    for idx, e in enumerate(data, start=1):
        fecha_formateada = ""
        if str(e['fecha_alta']) and str(e['fecha_alta']) != "None":
            try: fecha_formateada = datetime.strptime(str(e['fecha_alta']).split()[0], "%Y-%m-%d").strftime("%d/%m/%y")
            except: fecha_formateada = str(e['fecha_alta'])
        table_data.append([
            idx, Paragraph(e['referencia'], p_style_ref), Paragraph(e['organismo'], p_style), 
            Paragraph(e.get('ubicacion', ''), p_style), Paragraph(e['localidad'], p_style), 
            Paragraph(e['tipo_enlace'], p_style), e['ancho_banda'], 
            f"USD {format_currency(e['precio_venta_sin_iva'])}", 
            f"USD {format_currency(e['precio_venta_con_iva'])}", 
            f"USD {format_currency(e['costo_instalacion'])}", 
            f"ARS {format_currency(e['costo_mantenimiento_con_iva'])}", 
            fecha_formateada
        ])
    if not df.empty:
        v_si_sum = sum(float(x.get('precio_venta_sin_iva') or 0) for x in data)
        v_ci_sum = sum(float(x.get('precio_venta_con_iva') or 0) for x in data)
        inst_sum = sum(float(x.get('costo_instalacion') or 0) for x in data)
        mant_sum = sum(float(x.get('costo_mantenimiento_con_iva') or 0) for x in data)
        table_data.append(['TOTALES', '', '', '', '', '', '', f"USD {format_currency(v_si_sum)}", f"USD {format_currency(v_ci_sum)}", f"USD {format_currency(inst_sum)}", f"ARS {format_currency(mant_sum)}", ''])
        
    t = Table(table_data, colWidths=[18, 68, 110, 110, 75, 75, 45, 65, 65, 65, 65, 40], repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F497D')), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke), ('GRID', (0,0), (-1,-1), 0.5, colors.black), ('FONTSIZE', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 3), ('TOPPADDING', (0,0), (-1,-1), 3), ('LEFTPADDING', (0,0), (-1,-1), 2), ('RIGHTPADDING', (0,0), (-1,-1), 2), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (0,0), (-1,-1), 'CENTER')]))
    doc.build([Paragraph("Reporte de Facturación de Enlaces", getSampleStyleSheet()['Heading2']), Spacer(1, 10), t])
    
    db.add(ReporteFacturacion(archivo=filename, usuario=user, comentario=f.comentario))
    db.commit()
    return {"ok": True, "archivo": filename}

@app.delete("/facturaciones/{id}")
def eliminar_facturacion(id: int, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    rep = db.query(ReporteFacturacion).filter(ReporteFacturacion.id == id).first()
    if not rep: raise HTTPException(status_code=404, detail="Reporte no encontrado")
    if os.path.exists(os.path.join("adjuntos", rep.archivo)):
        try: os.remove(os.path.join("adjuntos", rep.archivo))
        except: pass
    db.delete(rep); db.commit()
    return {"ok": True}

@app.get("/reportes/excel")
def exportar_excel(organismo: Optional[str] = None, localidad: Optional[str] = None, db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    query = db.query(Enlace)
    if organismo: query = query.filter(Enlace.organismo.ilike(f"%{organismo}%"))
    if localidad: query = query.filter(Enlace.localidad.ilike(f"%{localidad}%"))
    data = [enlace_to_dict(e) for e in query.all() if e.eliminado not in [True, 1, "1", "true", "True"]]
            
    df = pd.DataFrame(data)
    wb = Workbook(); ws = wb.active
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.append(['IT', 'Referencia', 'Organismo', 'Ubicación', 'Localidad', 'Tipo', 'Ancho Banda', 'Costo Mensual', 'Vta s/IVA', 'Vta c/IVA', 'Instalacion', 'Mant.', 'Alta', 'Margen', 'Cupo', 'GPS', 'S/N Antena', 'S/N Modem', 'Nro TE', 'Nro ITEM'])
    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font; cell.border = border
    
    for idx, e in enumerate(data, start=1):
        fecha_formateada = ""
        if str(e['fecha_alta']) and str(e['fecha_alta']) != "None":
            try: fecha_formateada = datetime.strptime(str(e['fecha_alta']).split()[0], "%Y-%m-%d").strftime("%d/%m/%y")
            except: fecha_formateada = str(e['fecha_alta'])

        costo_mensual = float(e['precio_costo']) if e.get('precio_costo') else 0.0
        vta_c_iva = float(e['precio_venta_con_iva']) if e.get('precio_venta_con_iva') else 0.0

        row = [
            idx, e['referencia'], e['organismo'], e.get('ubicacion', ''), e['localidad'], e['tipo_enlace'], e['ancho_banda'], 
            f"USD {format_currency(costo_mensual)}", 
            f"USD {format_currency(e['precio_venta_sin_iva'])}", 
            f"USD {format_currency(vta_c_iva)}", 
            f"USD {format_currency(e['costo_instalacion'])}", 
            f"ARS {format_currency(e['costo_mantenimiento_con_iva'])}", 
            fecha_formateada, 
            f"USD {format_currency(vta_c_iva - costo_mensual)}",
            e.get('cupo_transferencia', ''), e.get('coordenadas_gps', ''), e.get('sn_antena', ''), e.get('sn_modem', ''), e.get('nro_te', ''), e.get('nro_item', '')
        ]
        ws.append(row)
        for cell in ws[ws.max_row]: cell.border = border
        
    if not df.empty:
        c_m_sum = sum(float(x.get('precio_costo') or 0) for x in data)
        v_si_sum = sum(float(x.get('precio_venta_sin_iva') or 0) for x in data)
        v_ci_sum = sum(float(x.get('precio_venta_con_iva') or 0) for x in data)
        inst_sum = sum(float(x.get('costo_instalacion') or 0) for x in data)
        mant_sum = sum(float(x.get('costo_mantenimiento_con_iva') or 0) for x in data)
        ws.append(['TOTALES', '', '', '', '', '', '', f"USD {format_currency(c_m_sum)}", f"USD {format_currency(v_si_sum)}", f"USD {format_currency(v_ci_sum)}", f"USD {format_currency(inst_sum)}", f"ARS {format_currency(mant_sum)}", '', f"USD {format_currency(v_ci_sum - c_m_sum)}", '', '', '', '', '', ''])
        
    stream = io.BytesIO(); wb.save(stream); stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=Reporte_Filtrado.xlsx"})


@app.get("/reportes/pdf")
def exportar_pdf(db: Session = Depends(get_db), user: str = Depends(verificar_usuario)):
    data = [enlace_to_dict(e) for e in db.query(Enlace).all() if e.eliminado not in [True, 1, "1", "true", "True"]]
    df = pd.DataFrame(data)
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), leftMargin=15, rightMargin=15, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    p_style = ParagraphStyle('WrapStyle', parent=styles['Normal'], fontSize=6, leading=7, alignment=TA_CENTER)
    p_style_ref = ParagraphStyle('RefStyle', parent=styles['Normal'], fontSize=5, leading=6, alignment=TA_CENTER)
    
    table_data = [['IT', 'Ref', 'Organismo', 'Ubicación', 'Localidad', 'Tipo', 'Ancho B.', 'Vta s/IVA', 'Vta c/IVA', 'Instal.', 'Mant.', 'Alta']]
    for idx, e in enumerate(data, start=1):
        fecha_formateada = ""
        if str(e['fecha_alta']) and str(e['fecha_alta']) != "None":
            try: fecha_formateada = datetime.strptime(str(e['fecha_alta']).split()[0], "%Y-%m-%d").strftime("%d/%m/%y")
            except: fecha_formateada = str(e['fecha_alta'])
        table_data.append([
            idx, Paragraph(e['referencia'], p_style_ref), Paragraph(e['organismo'], p_style), 
            Paragraph(e.get('ubicacion', ''), p_style), Paragraph(e['localidad'], p_style), 
            Paragraph(e['tipo_enlace'], p_style), e['ancho_banda'], 
            f"USD {format_currency(e['precio_venta_sin_iva'])}", f"USD {format_currency(e['precio_venta_con_iva'])}", 
            f"USD {format_currency(e['costo_instalacion'])}", f"ARS {format_currency(e['costo_mantenimiento_con_iva'])}", fecha_formateada
        ])
    if not df.empty:
        v_si_sum = sum(float(x.get('precio_venta_sin_iva') or 0) for x in data)
        v_ci_sum = sum(float(x.get('precio_venta_con_iva') or 0) for x in data)
        inst_sum = sum(float(x.get('costo_instalacion') or 0) for x in data)
        mant_sum = sum(float(x.get('costo_mantenimiento_con_iva') or 0) for x in data)
        table_data.append(['TOTALES', '', '', '', '', '', '', f"USD {format_currency(v_si_sum)}", f"USD {format_currency(v_ci_sum)}", f"USD {format_currency(inst_sum)}", f"ARS {format_currency(mant_sum)}", ''])
        
    t = Table(table_data, colWidths=[18, 68, 110, 110, 75, 75, 45, 65, 65, 65, 65, 40], repeatRows=1)
    t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F497D')), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke), ('GRID', (0,0), (-1,-1), 0.5, colors.black), ('FONTSIZE', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 3), ('TOPPADDING', (0,0), (-1,-1), 3), ('LEFTPADDING', (0,0), (-1,-1), 2), ('RIGHTPADDING', (0,0), (-1,-1), 2), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('ALIGN', (0,0), (-1,-1), 'CENTER')]))
    doc.build([Paragraph("Reporte Institucional de Enlaces", getSampleStyleSheet()['Heading2']), Spacer(1, 10), t])
    pdf_buffer.seek(0)
    return StreamingResponse(pdf_buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=Reporte.pdf"})
